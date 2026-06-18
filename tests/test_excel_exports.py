from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook

from apps.cadastrocomissao.service.excel_exporter import ExcelExporter
from apps.cadastromapa.servicos import (
    COLUNAS_EXPORTACAO,
    extrair_dados_pagina,
    gerar_planilha_cadastro_mapa,
)


EXPECTED_HEADERS = [
    "CPF/CNPJ",
    "NOME / RAZÃO SOCIAL",
    "FILIAL",
    "NÚMERO CONTRATO",
    "DATA CONTRATO",
    "PLANO",
    "TIPO DE PRODUTO",
    "OBSERVAÇÃO  CONTRATO",
    "PARCELA",
    "VENCIMENTO",
    "VALOR",
    "OBSERVAÇÃO PARCELA",
    "TEL. RESIDENCIAL 1",
    "TEL. RESIDENCIAL 2",
    "TEL. COMERCIAL 1",
    "TEL. COMERCIAL 2",
    "TEL. CELULAR 1",
    "TEL. CELULAR 2",
    "TEL. REFERÊNCIA 1",
    "OBS.  TEL. REFERÊNCIA 1",
    "TEL. REFERÊNCIA 2",
    "OBS.  TEL. REFERÊNCIA 2",
    "TEL. REFERÊNCIA 3",
    "OBS.  TEL. REFERÊNCIA 3",
    "EMAIL 1",
    "EMAIL 2",
    "EMAIL 3",
    "ENDEREÇO RES.",
    "NUMERO RES.",
    "COMPLEMENTO RES.",
    "BAIRRO RES.",
    "CEP  RES.",
    "CIDADE RES.",
    "UF RES.",
    "ENDEREÇO COM.",
    "NUMERO COM.",
    "COMPLEMENTO COM.",
    "BAIRRO COM.",
    "CEP  COM.",
    "CIDADE COM.",
    "UF COM.",
    "RG/IE",
    "DATA NASC.",
    "PAI",
    "MAE",
]


def row_values(sheet, row_number):
    return [sheet.cell(row_number, column).value for column in range(1, sheet.max_column + 1)]


def values_by_header(sheet, row_number):
    headers = row_values(sheet, 1)
    values = row_values(sheet, row_number)
    return dict(zip(headers, values))


def assert_import_headers(sheet):
    headers = row_values(sheet, 1)
    assert sheet.max_column == 45
    assert headers == EXPECTED_HEADERS


def assert_commercial_and_rg_fields_are_blank(row):
    for header in [
        "ENDEREÇO COM.",
        "NUMERO COM.",
        "COMPLEMENTO COM.",
        "BAIRRO COM.",
        "CEP  COM.",
        "CIDADE COM.",
        "UF COM.",
        "RG/IE",
    ]:
        assert row[header] in (None, "")


def assert_header_style_without_fill(sheet):
    for cell in sheet[1]:
        assert cell.font.bold is True
        assert cell.fill.fill_type is None
        assert cell.border.left.style is None
        assert cell.border.right.style is None
        assert cell.border.top.style is None
        assert cell.border.bottom.style is None
        assert cell.alignment.horizontal is None
        assert cell.alignment.vertical is None


def assert_without_auto_filter(sheet):
    assert sheet.auto_filter.ref is None
    assert sheet.freeze_panes is None
    assert list(sheet.column_dimensions) == []


def test_cadastrocomissao_export_has_expected_columns_and_values(tmp_path):
    extraction = {
        "tables": [
            {
                "id": "buyers",
                "rows": [
                    {
                        "Documento": "123.456.789-00",
                        "Comprador": "Cliente Comissao",
                        "Leilao": "Virtual Teste - 09/06/2026",
                        "Data do Leilao": "09/06/2026",
                        "Observação": "Comissao de compra - Teste",
                        "Vencimento": "24/06/2026",
                        "Valor": "R$ 1.234,50",
                        "Telefone 1": "(11) 99999-9999",
                        "Telefone 2": "(11) 3333-4444",
                        "Telefone 3": "(11) 88888-7777",
                        "Email 1": "cliente@example.com",
                        "Email 2": "financeiro@example.com",
                        "Email 3": "extra@example.com",
                        "Endereco": "Rua A, 123 - Bloco B - Sao Paulo - SP - CEP: 01000-000",
                        "Bairro": "Centro",
                        "CEP": "01000-000",
                        "Cidade": "Sao Paulo",
                        "UF": "SP",
                    }
                ],
            }
        ]
    }
    selection = {
        "tables": [{"id": "buyers", "include": True}],
        "fields": [],
        "custom_columns": [],
        "options": {"exclude_missing_document": False},
    }

    output_path = ExcelExporter(tmp_path).export(extraction, selection, id_cadastro=4)
    workbook = load_workbook(output_path)
    sheet = workbook["Sheet1"]
    row = values_by_header(sheet, 2)

    assert workbook.sheetnames == ["Sheet1"]
    assert_import_headers(sheet)
    assert_without_auto_filter(sheet)
    assert sheet.max_row == 2
    assert row["CPF/CNPJ"] == "123.456.789-00"
    assert row["NOME / RAZÃO SOCIAL"] == "Cliente Comissao"
    assert row["NÚMERO CONTRATO"] == f"4{datetime.now().strftime('%d%m%y')}001_Teste-09/06/2026"
    assert row["DATA CONTRATO"] == "09/06/2026"
    assert row["TIPO DE PRODUTO"] == "Duplicata"
    assert row["OBSERVAÇÃO  CONTRATO"] == "Comissao de compra - Teste"
    assert row["PARCELA"] == 1
    assert row["VENCIMENTO"] == "24/06/2026"
    assert row["VALOR"] == "1234,5"
    assert row["TEL. RESIDENCIAL 1"] == "1199999-9999"
    assert row["TEL. RESIDENCIAL 2"] == "113333-4444"
    assert row["TEL. CELULAR 1"] == "1188888-7777"
    assert row["EMAIL 1"] == "cliente@example.com"
    assert row["EMAIL 2"] == "financeiro@example.com"
    assert row["EMAIL 3"] == "extra@example.com"
    assert row["ENDEREÇO RES."] == "Rua A"
    assert row["NUMERO RES."] == "123"
    assert row["COMPLEMENTO RES."] == "Bloco B"
    assert row["BAIRRO RES."] == "Centro"
    assert row["CEP  RES."] == "01000000"
    assert row["CIDADE RES."] == "Sao Paulo"
    assert row["UF RES."] == "SP"
    assert_commercial_and_rg_fields_are_blank(row)


def test_cadastrocomissao_export_skips_missing_document_when_option_is_enabled(tmp_path):
    extraction = {
        "tables": [
            {
                "id": "buyers",
                "rows": [
                    {"Documento": "", "Comprador": "Sem Documento"},
                    {"Documento": "987.654.321-00", "Comprador": "Com Documento"},
                ],
            }
        ]
    }

    output_path = ExcelExporter(tmp_path).export(
        extraction,
        {
            "tables": [{"id": "buyers", "include": True}],
            "fields": [],
            "custom_columns": [],
            "options": {"exclude_missing_document": True},
        },
    )
    sheet = load_workbook(output_path)["Sheet1"]
    row = values_by_header(sheet, 2)

    assert_import_headers(sheet)
    assert_without_auto_filter(sheet)
    assert sheet.max_row == 2
    assert row["CPF/CNPJ"] == "987.654.321-00"
    assert row["NOME / RAZÃO SOCIAL"] == "Com Documento"


def test_cadastrocomissao_contract_uses_9_when_registration_id_is_missing(tmp_path):
    extraction = {
        "tables": [
            {
                "id": "buyers",
                "rows": [
                    {
                        "Documento": "123.456.789-00",
                        "Comprador": "Cliente",
                        "Leilao": "Virtual Leilao Teste - 01/01/2026",
                        "Data do Leilao": "01/01/2026",
                    }
                ],
            }
        ]
    }
    output_path = ExcelExporter(tmp_path).export(
        extraction,
        {"tables": [{"id": "buyers", "include": True}], "fields": [], "custom_columns": [], "options": {}},
    )
    row = values_by_header(load_workbook(output_path)["Sheet1"], 2)

    assert row["NÚMERO CONTRATO"].startswith(f"9{datetime.now().strftime('%d%m%y')}001_")


def test_cadastromapa_export_has_expected_columns_values_and_one_row_per_installment(tmp_path):
    paginas = [
        {
            "leilao": "Leilao Mapa - 09/06/2026",
            "datacontrato": "09/06/2026",
            "produto": "Duplicata",
            "comprador": "Cliente Mapa",
            "cpfcnpj": "12.345.678/0001-99",
            "endereco": "Avenida B",
            "numero": "456",
            "bairro": "Jardim",
            "cep": "12345-678",
            "cidade": "Ribeirao Preto",
            "uf": "SP",
            "complemento": "Sala 2",
            "telefones": ["(16) 99999-0000", "(16) 3333-2222", "(16) 98888-7777"],
            "emails": ["mapa@example.com", "boleto@example.com", "extra@example.com"],
            "observacao": "Leilao Mapa - Lote(s): 10",
            "parcelas": [
                {"parcela": 1, "vencimento": "20/06/2026", "valor": 1500.75},
                {"parcela": 2, "vencimento": "20/07/2026", "valor": 2500.0},
            ],
        }
    ]

    output_path = Path(tmp_path) / "cadastromapa.xlsx"
    output_path.write_bytes(gerar_planilha_cadastro_mapa(paginas, id_cadastro=7).getvalue())
    workbook = load_workbook(output_path)
    sheet = workbook.active
    first_row = values_by_header(sheet, 2)
    second_row = values_by_header(sheet, 3)

    assert_import_headers(sheet)
    assert_without_auto_filter(sheet)
    assert sheet.max_row == 3
    assert first_row["CPF/CNPJ"] == "12.345.678/0001-99"
    assert first_row["NOME / RAZÃO SOCIAL"] == "Cliente Mapa"
    assert first_row["NÚMERO CONTRATO"] == f"7{datetime.now().strftime('%d%m%y')}001_Leilao Mapa - 09/06/2026"
    assert first_row["DATA CONTRATO"] == "09/06/2026"
    assert first_row["TIPO DE PRODUTO"] == "Duplicata"
    assert first_row["OBSERVAÇÃO  CONTRATO"] == "Leilao Mapa - Lote(s): 10"
    assert first_row["PARCELA"] == 1
    assert first_row["VENCIMENTO"] == "20/06/2026"
    assert first_row["VALOR"] == 1500.75
    assert first_row["TEL. RESIDENCIAL 1"] == "1699999-0000"
    assert first_row["TEL. RESIDENCIAL 2"] == "163333-2222"
    assert first_row["TEL. CELULAR 1"] == "1698888-7777"
    assert first_row["EMAIL 1"] == "mapa@example.com"
    assert first_row["EMAIL 2"] == "boleto@example.com"
    assert first_row["EMAIL 3"] == "extra@example.com"
    assert first_row["ENDEREÇO RES."] == "Avenida B"
    assert first_row["NUMERO RES."] == "456"
    assert first_row["COMPLEMENTO RES."] == "Sala 2"
    assert first_row["BAIRRO RES."] == "Jardim"
    assert first_row["CEP  RES."] == "12345678"
    assert first_row["CIDADE RES."] == "Ribeirao Preto"
    assert first_row["UF RES."] == "SP"
    assert_commercial_and_rg_fields_are_blank(first_row)
    assert second_row["PARCELA"] == 2
    assert second_row["VENCIMENTO"] == "20/07/2026"
    assert second_row["VALOR"] == 2500


def test_cadastromapa_pdf_extraction_adds_animal_to_observation():
    texto_pdf = """
    Nelore Pintado Rubro Negro - Prenhezes, Aspirações e Babys - 15/05/2026
    Cadastro de Compradores e Dados para Emissão de Nota Fiscal
    Birigui - SP
    Vendedor: Cristiano Freitas de Oliveira
    Comprador: Beatriz Aparecida de Souza Guimarães - 039.260.886-30 - Data nascimento: 19/03/1956
    Endereço: Avenida Jacinto Barbosa, 314 - São Francisco - Patrocinio - MG - CEP : 38742-008
    E-mail: marcos@aguimaraes.com.br
    Qtde
    Lote
    Sexo
    Animal
    Crias/Recep
    Condição de Pagamento Desc %
    Total R$
    Total Líquido R$
    Peso M.
    R$/KG
    1
    15
    F
    (6,25%)  Nelore Pintado PO CFO 110 (4m)
    01+29 PARCELAS IGUAIS
    MENSAL
    0,00
    7.500,00
    7.500,00
    170
    R$
    352,94
    01/ 30 - 15/05/2026 - R$ 250,00
    """

    pagina = extrair_dados_pagina(1, texto_pdf)

    assert (
        pagina["observacao"]
        == "Nelore Pintado Rubro Negro - Prenhezes, Aspirações e Babys - 15/05/2026 - "
        "Lote: 15 - (6,25%)  Nelore Pintado PO CFO 110 (4m)"
    )


def test_cadastromapa_observation_accepts_alphanumeric_lot():
    texto_pdf = """
    Nelore Gibertoni Gerações - Evoluzione - 10/05/2026
    Cadastro de Compradores e Dados para Emissão de Nota Fiscal
    Vendedor: Washington Dias Janota Antunes
    Comprador: Alexandre Lima Sangali - 011.993.365-99
    Endereço: Waldemar Falcao, 999 - Horto Florestal - Salvador - BA - CEP : 40295-010
    E-mail: alexandre@example.com
    Qtde
    Lote
    Sexo
    Animal
    Crias/Recep
    Condição de Pagamento Desc %
    Total R$
    Total Líquido R$
    Peso M.
    R$/KG
    1
    69A
    S
    (50,00%)  Nelore PO
    01+29 PARCELAS IGUAIS
    MENSAL
    0,00
    150.000,00
    150.000,00
    0
    R$ 0,00
    """

    pagina = extrair_dados_pagina(1, texto_pdf)

    assert (
        pagina["observacao"]
        == "Nelore Gibertoni Gerações - Evoluzione - 10/05/2026 - "
        "Lote: 69A - (50,00%)  Nelore PO"
    )


def test_cadastromapa_observation_handles_multiple_items_and_wrapped_lot():
    texto_pdf = """
    Tomorrow Farm - Grupo Mônica 2026 - 10/04/2026
    Cadastro de Compradores e Dados para Emissão de Nota Fiscal
    Vendedor: Talita Taynara Marques Gonçalves
    Comprador: Cliente Teste - 07.589.548/0009-15
    Endereço: Avenida 21 de Abril, 18 - Iaciara - GO - CEP : 73920-000
    E-mail: teste@example.com
    Qtde
    Lote
    Sexo
    Animal
    Crias/Recep
    Condição de Pagamento Desc %
    Total R$
    Total Líquido R$
    Peso M.
    R$/KG
    1
    10
    S
    (50,00%)  Nelore PO
    01+29 PARCELAS IGUAIS
    MENSAL
    0,00
    630.000,00
    630.000,00
    0
    R$ 0,00
    1
    MONIC
    A
    F
    (50,00%)  Nelore PO LILL 2666 (47m)
    01+29 PARCELAS IGUAIS
    MENSAL
    0,00
    7.620.000,00
    7.620.000,00
    0
    R$ 0,00
    Machos
    """

    pagina = extrair_dados_pagina(1, texto_pdf)

    assert (
        pagina["observacao"]
        == "Tomorrow Farm - Grupo Mônica 2026 - 10/04/2026 - "
        "Lote: 10 - (50,00%)  Nelore PO; "
        "Lote: MONICA - (50,00%)  Nelore PO LILL 2666 (47m)"
    )


def test_export_headers_are_bold_and_without_fill(tmp_path):
    comissao_path = ExcelExporter(tmp_path).export(
        {"metadata": {"filename": "teste.pdf", "pages": 1}, "fields": [], "tables": [], "pages": []},
        {"tables": [], "fields": [], "custom_columns": [], "options": {}},
    )
    comissao_sheet = load_workbook(comissao_path).active

    mapa_path = Path(tmp_path) / "mapa.xlsx"
    mapa_path.write_bytes(
        gerar_planilha_cadastro_mapa(
            [
                {
                    "cpfcnpj": "123",
                    "comprador": "Cliente",
                    "parcelas": [{"parcela": 1, "vencimento": "10/06/2026", "valor": 1}],
                }
            ]
        ).getvalue()
    )
    mapa_sheet = load_workbook(mapa_path).active

    assert_header_style_without_fill(comissao_sheet)
    assert_header_style_without_fill(mapa_sheet)
