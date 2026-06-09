import re
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from uuid import uuid4

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


class ExcelExporter:
    HEADER_FILL = PatternFill(fill_type=None)
    HEADER_FONT = Font(bold=True)
    SUBTLE_FILL = PatternFill("solid", fgColor="D9EAF7")
    MISSING_FILL = PatternFill("solid", fgColor="FDE2E1")
    MISSING_FONT = Font(color="991B1B", bold=True)
    IMPORT_HEADERS = [
        "CPF/CNPJ",
        "NOME / RAZÃO SOCIAL",
        "FILIAL",
        "NÚMERO CONTRATO",
        "DATA CONTRATO",
        "PLANO",
        "TIPO DE PRODUTO",
        "OBSERVAÇÃO  CONTRATO",
        "PARCELA",
        "VENCIMENTO",
        "VALOR",
        "OBSERVAÇÃO PARCELA",
        "TEL. RESIDENCIAL 1",
        "TEL. RESIDENCIAL 2",
        "TEL. COMERCIAL 1",
        "TEL. COMERCIAL 2",
        "TEL. CELULAR 1",
        "TEL. CELULAR 2",
        "TEL. REFERÊNCIA 1",
        "OBS.  TEL. REFERÊNCIA 1",
        "TEL. REFERÊNCIA 2",
        "OBS.  TEL. REFERÊNCIA 2",
        "TEL. REFERÊNCIA 3",
        "OBS.  TEL. REFERÊNCIA 3",
        "EMAIL 1",
        "EMAIL 2",
        "EMAIL 3",
        "ENDEREÇO RES.",
        "NUMERO RES.",
        "COMPLEMENTO RES.",
        "BAIRRO RES.",
        "CEP  RES.",
        "CIDADE RES.",
        "UF RES.",
        "ENDEREÇO COM.",
        "NUMERO COM.",
        "COMPLEMENTO COM.",
        "BAIRRO COM.",
        "CEP  COM.",
        "CIDADE COM.",
        "UF COM.",
        "RG/IE",
        "DATA NASC.",
        "PAI",
        "MAE",
    ]

    def __init__(self, export_folder):
        self.export_folder = Path(export_folder)
        self.export_folder.mkdir(parents=True, exist_ok=True)

    def export(self, extraction, selection, id_cadastro=None):
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        self.id_cadastro = self._normalize_registration_id(id_cadastro)

        selected_tables_count, wrote_import_sheet = self._write_selected_tables(workbook, extraction, selection)
        if not wrote_import_sheet:
            buyer_table = self._find_buyer_table(extraction)
            if buyer_table:
                self._write_buyers_import_sheet(workbook, buyer_table, selection.get("options", {}))
                selected_tables_count += 1
                wrote_import_sheet = True

        if not wrote_import_sheet:
            self._write_selected_fields(
                workbook,
                extraction,
                selection,
                force_fallback=selected_tables_count == 0,
            )
        if not wrote_import_sheet:
            self._write_source_text(workbook, extraction)

        filename = f"exportacao_pdf_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid4().hex[:8]}.xlsx"
        output_path = self.export_folder / filename
        workbook.save(output_path)
        return output_path

    def _write_selected_fields(self, workbook, extraction, selection, force_fallback=True):
        selected_fields = selection.get("fields", [])
        custom_columns = selection.get("custom_columns", [])
        fields_by_key = {field["key"]: field for field in extraction.get("fields", [])}

        columns = []
        values = []

        for selected in selected_fields:
            if not selected.get("include"):
                continue
            field = fields_by_key.get(selected.get("key"))
            if not field:
                continue
            columns.append(selected.get("label") or field["label"])
            values.append(field.get("value", ""))

        for custom in custom_columns:
            name = (custom.get("label") or "").strip()
            if name:
                columns.append(name)
                values.append(self._resolve_custom_value(extraction, custom))

        if not columns and not force_fallback:
            return False

        if not columns:
            columns = ["Arquivo", "Paginas", "Campos detectados"]
            metadata = extraction.get("metadata", {})
            values = [
                metadata.get("filename", ""),
                metadata.get("pages", 0),
                len(extraction.get("fields", [])),
            ]

        sheet = workbook.create_sheet("Campos selecionados")
        sheet.append(columns)
        sheet.append(values)
        self._style_table(sheet, header_row=1)
        return True

    def _write_selected_tables(self, workbook, extraction, selection):
        selected_tables = selection.get("tables", [])
        tables_by_id = {table["id"]: table for table in extraction.get("tables", [])}
        written_count = 0
        wrote_import_sheet = False

        selected_buyer = next(
            (
                selected
                for selected in selected_tables
                if selected.get("include") and selected.get("id") == "buyers"
            ),
            None,
        )
        buyer_table = tables_by_id.get("buyers") if selected_buyer else None
        if buyer_table:
            self._write_buyers_import_sheet(workbook, buyer_table, selection.get("options", {}))
            return 1, True

        for selected in selected_tables:
            if not selected.get("include"):
                continue

            table = tables_by_id.get(selected.get("id"))
            if not table:
                continue

            selected_columns = [column for column in selected.get("columns", []) if column.get("include")]
            if not selected_columns:
                continue

            title = selected.get("title") or table.get("title") or "Tabela"
            sheet = workbook.create_sheet(self._safe_sheet_name(title))
            sheet.append([column.get("label") or column.get("key") for column in selected_columns])

            for row in table.get("rows", []):
                sheet.append([
                    self._display_table_value(row, column.get("key"))
                    for column in selected_columns
                ])

            self._style_table(sheet, header_row=1)
            self._style_missing_cells(sheet, table, selected_columns)
            written_count += 1

        return written_count, wrote_import_sheet

    def _write_buyers_import_sheet(self, workbook, table, options=None):
        sheet = workbook.create_sheet("Sheet1")
        sheet.append(self.IMPORT_HEADERS)

        rows = table.get("rows", [])
        if (options or {}).get("exclude_missing_document"):
            rows = [row for row in rows if not self._is_blank_document(row.get("Documento", ""))]

        for index, row in enumerate(rows, start=1):
            sheet.append(self._build_import_row(row, index))

        self._style_table(sheet, header_row=1)

    def _build_import_row(self, row, index):
        address = self._split_address(row)
        phones = [
            row.get("Telefone 1", ""),
            row.get("Telefone 2", ""),
            row.get("Telefone 3", ""),
        ]

        return [
            row.get("Documento", ""),
            row.get("Comprador", ""),
            "",
            self._build_contract_number(row, index),
            row.get("Data do Leilao", ""),
            "",
            "Duplicata",
            row.get("Observação", ""),
            1,
            self._export_due_date(row),
            self._normalize_money(row.get("Valor", "")),
            "",
            self._list_value(phones, 0),
            self._list_value(phones, 1),
            "",
            "",
            self._list_value(phones, 2),
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            row.get("Email 1", ""),
            row.get("Email 2", ""),
            row.get("Email 3", ""),
            address["street"],
            address["number"],
            address["complement"],
            address["neighborhood"],
            self._digits_only(row.get("CEP", "")),
            row.get("Cidade", ""),
            row.get("UF", ""),
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ]

    def _build_contract_number(self, row, index):
        run_prefix = datetime.now().strftime(f"{self.id_cadastro}%d%m%Y")
        auction_name = self._contract_auction_name(row)
        return f"{run_prefix}{index:03d}_{auction_name}"

    def _contract_auction_name(self, row):
        auction_label = self._clean_scalar(row.get("Leilao", ""))
        auction_date = self._clean_scalar(row.get("Data do Leilao", ""))
        label_without_date = auction_label

        if auction_date:
            label_without_date = re.sub(
                rf"\s*-\s*{re.escape(auction_date)}\s*$",
                "",
                label_without_date,
            ).strip()

        label_without_date = re.sub(r"^Virtual\s+", "", label_without_date, flags=re.IGNORECASE).strip()
        label_without_date = re.sub(r"\s+-\s+", "-", label_without_date)

        if auction_date and auction_date in label_without_date:
            return label_without_date
        if auction_date:
            return f"{label_without_date}-{auction_date}"
        return label_without_date or auction_label

    def _split_address(self, row):
        address = self._clean_scalar(row.get("Endereco", ""))
        city = self._clean_scalar(row.get("Cidade", ""))
        uf = self._clean_scalar(row.get("UF", ""))
        cep = self._clean_scalar(row.get("CEP", ""))

        base = address
        if cep:
            base = re.sub(r"\s+-\s*CEP\s*:?\s*" + re.escape(cep) + r"\s*$", "", base, flags=re.IGNORECASE)
        if city and uf:
            base = re.sub(
                r"\s+-\s*" + re.escape(city) + r"\s+-\s*" + re.escape(uf) + r"\s*$",
                "",
                base,
                flags=re.IGNORECASE,
            )

        parts = [part.strip(" .") for part in re.split(r"\s+-\s+", base) if part.strip(" .")]
        first_part = parts[0] if parts else base
        complement = " - ".join(parts[1:])
        street, number = self._split_street_number(first_part)

        return {
            "street": street,
            "number": number,
            "complement": complement,
            "neighborhood": self._clean_scalar(row.get("Bairro", "")) or "Centro",
        }

    def _split_street_number(self, value):
        text = self._clean_scalar(value)
        if not text:
            return "", ""

        comma_parts = [part.strip() for part in text.split(",") if part.strip()]
        if len(comma_parts) >= 2:
            number = comma_parts[-1]
            street = ", ".join(comma_parts[:-1])
            return street, number

        match = re.match(r"^(?P<street>.+?)\s+(?P<number>S/?N|SN|S\.N\.|N/?A|\d+[A-Za-z]?)$", text, flags=re.IGNORECASE)
        if match:
            return match.group("street").strip(" ,"), match.group("number").strip()

        return text, ""

    def _normalize_money(self, value):
        text = self._clean_scalar(value)
        if not text:
            return ""

        normalized = text.replace("R$", "").replace(".", "").replace(" ", "").replace(",", ".")
        try:
            amount = Decimal(normalized)
        except InvalidOperation:
            return text.replace("R$", "").strip()

        formatted = f"{amount:.2f}".replace(".", ",")
        return formatted.rstrip("0").rstrip(",")

    def _export_due_date(self, row):
        due_date = self._clean_scalar(row.get("Vencimento", ""))
        auction_date = self._clean_scalar(row.get("Data do Leilao", ""))
        if due_date and due_date != auction_date:
            return due_date

        return (datetime.now() + timedelta(days=15)).strftime("%d/%m/%Y")

    def _digits_only(self, value):
        return re.sub(r"\D+", "", self._clean_scalar(value))

    def _is_blank_document(self, value):
        return self._clean_scalar(value).lower() in {"", "-", "none", "null", "pendente"}

    def _normalize_registration_id(self, value):
        value = self._clean_scalar(value)
        return value if re.fullmatch(r"\d", value) else "9"

    def _list_value(self, values, index):
        return values[index] if len(values) > index else ""

    def _clean_scalar(self, value):
        value = "" if value is None else str(value)
        value = value.replace("\x00", " ")
        return re.sub(r"\s+", " ", value).strip()

    def _write_source_text(self, workbook, extraction):
        sheet = workbook.create_sheet("Texto extraido")
        sheet.append(["Pagina", "Texto"])

        for page in extraction.get("pages", []):
            sheet.append([page.get("number", ""), page.get("text", "")])

        self._style_table(sheet, header_row=1)

    def _style_table(self, sheet, header_row):
        for cell in sheet[header_row]:
            cell.font = self.HEADER_FONT

    def _safe_sheet_name(self, value):
        invalid_chars = ["\\", "/", "*", "[", "]", ":", "?"]
        for char in invalid_chars:
            value = value.replace(char, " ")
        value = " ".join(value.split()) or "Tabela"
        return value[:31]

    def _display_table_value(self, row, column_key):
        if self._is_missing_cell(row, column_key):
            return "PENDENTE"
        return row.get(column_key, "")

    def _style_missing_cells(self, sheet, table, selected_columns):
        if table.get("id") != "buyers":
            return

        for row_index, row in enumerate(table.get("rows", []), start=2):
            for column_index, column in enumerate(selected_columns, start=1):
                if self._is_missing_cell(row, column.get("key")):
                    cell = sheet.cell(row=row_index, column=column_index)
                    cell.fill = self.MISSING_FILL
                    cell.font = self.MISSING_FONT

    def _is_missing_cell(self, row, column_key):
        return column_key in set(row.get("__missing_fields", []))

    def _find_buyer_table(self, extraction):
        return next((table for table in extraction.get("tables", []) if table.get("id") == "buyers"), None)

    def _resolve_custom_value(self, extraction, custom):
        value = (custom.get("value") or "").strip()
        if custom.get("mode") != "regex" or not value:
            return value

        try:
            match = re.search(value, extraction.get("full_text", ""), flags=re.IGNORECASE | re.MULTILINE)
        except re.error:
            return ""

        if not match:
            return ""
        if match.groups():
            return match.group(1).strip()
        return match.group(0).strip()
